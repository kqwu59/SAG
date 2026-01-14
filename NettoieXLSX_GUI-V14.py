#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
NettoieXLSX_GUI.py (V14)
- Fichiers optionnels et ordre UI : Commandes, Constatations, Factures, Envoi BDC, Workflow
- Nettoyages :
  * Commandes : DROP 20 premières lignes, garder & ordonner :
      N° commande, Libellé, Fournisseur, Montant HT, Type de flux, Nature de dépense, Statut, Ind. Visa, Auteur
    Filtrer : Fournisseur == 'FCM 3MUNDI ESR-M' OU Nature de dépense == 'Mission'.
  * Constatations : DROP 17 premières lignes, garder :
      Commande, extrait commande (=LEFT(Commande,5)), Statut.
  * Envoi BDC : garder STRICTEMENT les 3 premières colonnes -> renommer :
      Commande, Date envoi, Agent.
  * Factures : DROP 19 premières lignes, garder :
      N° commande, Montant HT, Date de règlement.
    Filtrer : Nature de dépense == 'MI' OU Fournisseur == 'FCM 3MUNDI ESR-M'.
  * Workflow : garder tout (utilisé pour Global!H), lecture sous “Liste des résultats” si présent.
- Dates : suppression de l'heure dans toutes les feuilles écrites (format dd/mm/yyyy).
- Sortie (si présentes) : Commande, Envoi BDC, Constatation, Factures, Workflow, Global
- Global (A..K) :
    A BDC      = Commande.N° commande
    B Objet    = Commande.Libellé
    C Fourn.   = Commande.Fournisseur
    D HT       = Commande.Montant HT
    E Visa     = Commande.Ind. Visa
    F Envoyé   = texte(Date envoi, dd/mm/yyyy) + " " + Agent (jointure BDC=Commande)
    G SF       = 'ss objet Régul CA' si C == 'BNP PARIBAS - REGULARISATION CARTE ACHAT'
                 sinon si F contient 'ss objet Régul CA' -> Statut Constatation (BDC complet, sinon LEFT(BDC,5))
                 sinon 'Pas de SF connu'
    H Workflow = valeur workflow (date/état), jointure BDC, date-only si datetime
    I Payé     = 0 facture -> 'pas de paiement connu'
                 1 facture -> Date de règlement (date-only si possible, sinon valeur brute ou 'date manquante')
                 >=2 factures -> 'n paiements' (singulier/pluriel correct)
    J Solde    = D (HT commande) - somme(Montant HT factures du BDC) [numérique, 2 décimales]
    K Statut   = Commande.Statut
- Déduplication Global : suppression des doublons STRICTS (toutes colonnes A..K identiques). Si même BDC mais
  colonnes différentes, toutes les variantes sont conservées.
"""

import os
import unicodedata
import datetime as dt
from decimal import Decimal, InvalidOperation

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

# -------- Réglages d'affichage --------
GLOBAL_WIDTH_OFFSET = 0.64  # correction écart Excel
GLOBAL_COLUMN_WIDTHS = [7.09, 36.09, 70, 12.09, 16, 14, 16.82, 30, 8.09, 8.09, 12.0]  # A..K

GLOBAL_RULES_TEXT = (
    "Règles des colonnes Global (A→K)\n\n"
    "A • BDC : repris de Commandes.N° commande (après filtres : Mission / FCM 3MUNDI ESR-M exclus).\n"
    "B • OBJET : Commandes.Libellé.\n"
    "C • FOURN. : Commandes.Fournisseur.\n"
    "D • HT : Commandes.Montant HT.\n"
    "E • VISA : Commandes.Ind. Visa.\n"
    "F • ENVOYE : jointure Envoi BDC sur BDC=Commande → 'Date envoi (dd/mm/yyyy) + espace + Agent'.\n"
    "G • SF :\n"
    "    - si FOURN. = 'BNP PARIBAS - REGULARISATION CARTE ACHAT' → 'ss objet Régul CA'\n"
    "    - sinon, si ENVOYE contient 'ss objet Régul CA' → Constatation.Statut (recherche par BDC complet,\n"
    "      sinon par les 5 premiers caractères)\n"
    "    - sinon → 'Pas de SF connu'.\n"
    "H • WORKFLOW : valeur depuis Workflow (colonne 'Date' prioritaire, sinon 'Statut'), jointure sur BDC.\n"
    "I • PAYE :\n"
    "    - 0 facture pour le BDC → 'pas de paiement connu'\n"
    "    - 1 facture → afficher la Date de règlement (date-only si possible ; sinon valeur brute ; sinon 'date manquante')\n"
    "    - ≥2 factures → 'n paiements'\n"
    "J • SOLDE : HT (D) − somme des Montants HT de toutes les lignes Factures associées au BDC (2 décimales).\n"
    "K • STATUT : Commandes.Statut.\n\n"
    "Déduplication : si deux lignes Global (A..K) sont STRICTEMENT identiques, une seule est conservée.\n"
    "Si le même BDC présente des différences sur au moins une colonne, toutes les lignes sont gardées."
)

GLOBAL_COVER_TEXT = (
    "PAGE DE GARDE - Colonnes de l'onglet Global\n\n"
    "A • BDC : repris de Commandes.N° commande (après filtres : Mission / FCM 3MUNDI ESR-M exclus).\n"
    "B • OBJET : Commandes.Libellé.\n"
    "C • FOURN. : Commandes.Fournisseur.\n"
    "D • HT : Commandes.Montant HT.\n"
    "E • VISA : Commandes.Ind. Visa.\n"
    "F • ENVOYE : jointure Envoi BDC sur BDC=Commande → 'Date envoi (dd/mm/yyyy) + espace + Agent'.\n"
    "G • SF :\n"
    "    - si FOURN. = 'BNP PARIBAS - REGULARISATION CARTE ACHAT' → 'ss objet Régul CA'\n"
    "    - sinon, si ENVOYE contient 'ss objet Régul CA' → Constatation.Statut (recherche par BDC complet,\n"
    "      sinon par les 5 premiers caractères)\n"
    "    - sinon → 'Pas de SF connu'.\n"
    "H • WORKFLOW : valeur depuis Workflow (colonne 'Date' prioritaire, sinon 'Statut'), jointure sur BDC.\n"
    "I • PAYE :\n"
    "    - 0 facture pour le BDC → 'pas de paiement connu'\n"
    "    - 1 facture → afficher la Date de règlement (date-only si possible ; sinon valeur brute ; sinon 'date manquante')\n"
    "    - ≥2 factures → 'n paiements'\n"
    "J • SOLDE : HT (D) − somme des Montants HT de toutes les lignes Factures associées au BDC (2 décimales).\n"
    "K • STATUT : Commandes.Statut.\n\n"
    "Déduplication : si deux lignes Global (A..K) sont STRICTEMENT identiques, une seule est conservée.\n"
    "Si le même BDC présente des différences sur au moins une colonne, toutes les lignes sont gardées."
)

INTRO_LOG_TEXT = (
    "La présente macro permet d’avoir une vision globale du traitement des commandes de la base 0018.\n\n"
    "Pour ce faire, elle exploite plusieurs fichiers au format EXCEL :\n"
    "- 3 extractions sous Geslab :\n"
    "   « commandes / réservations » : RAJOUTER dans les paramètres d’affichage : Type de flux et Auteur\n"
    "   « constatation »\n"
    "   « facture »\n"
    "- l’extraction des « workflows » sous DMF\n"
    "- le fichier « Envoi BDC » sous SAG/TUTOS complété lors du traitement des bons de commande\n\n"
    "Pour garantir une bonne utilisation de la macro, mettez les bons fichiers sur la bonne ligne correspondante\n"
    "(Astuce : glissez vos fichiers .xlsx sur les champs !)\n\n"
    "Dans les fichiers extraits de Geslab, seules les lignes sous 'Liste des résultats' seront prises en compte\n"
)

# Drag & drop (optionnel)
try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    DND_AVAILABLE = True
except Exception:
    DND_AVAILABLE = False

# -------- Helpers --------
def strip_accents(text: str) -> str:
    if text is None: return ""
    text = str(text)
    text = unicodedata.normalize("NFD", text)
    return "".join(ch for ch in text if unicodedata.category(ch) != "Mn")

def normalize_colname(name: str) -> str:
    s = strip_accents(str(name)).lower()
    for ch in ["\n","\r","\t"]: s = s.replace(ch," ")
    s = "".join(c if c.isalnum() else " " for c in s)
    return " ".join(s.split())

def to_date_only(value):
    if pd.isna(value): return ""
    if isinstance(value, (pd.Timestamp, dt.datetime)): return value.date()
    if isinstance(value, dt.date): return value
    try:
        parsed = pd.to_datetime(value, dayfirst=True, errors="coerce")
        if pd.notna(parsed): return parsed.date()
    except Exception:
        pass
    return value

def date_to_text_dmy(value):
    d = to_date_only(value)
    if isinstance(d, dt.date): return d.strftime("%d/%m/%Y")
    return str(d).strip()

def strip_times_in_worksheet(ws):
    import re
    pat = re.compile(r"^\s*(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})\s+\d{1,2}:\d{2}(:\d{2})?\s*$")
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, (pd.Timestamp, dt.datetime)):
                cell.value = v.date()
                cell.number_format = "dd/mm/yyyy"
            elif isinstance(v, str) and pat.match(v):
                d = pd.to_datetime(v, dayfirst=True, errors="coerce")
                if pd.notna(d):
                    cell.value = d.date()
                    cell.number_format = "dd/mm/yyyy"

def read_after_skip(xlsx_path: str, skip_rows: int):
    df_all = pd.read_excel(xlsx_path, sheet_name=0, header=None, engine="openpyxl")
    df = df_all.iloc[skip_rows:].copy()
    # entête = première ligne avec ≥2 valeurs non vides
    header_rel_idx = None
    for i in range(len(df)):
        row = df.iloc[i]
        if row.notna().sum() >= 2:
            header_rel_idx = i
            break
    if header_rel_idx is None:
        return pd.DataFrame()
    header_row = df.iloc[header_rel_idx].astype(str).str.strip().tolist()
    data = df.iloc[header_rel_idx+1:].copy()
    data.columns = header_row
    data = data.loc[:, data.columns.notnull()]
    data = data.dropna(how='all')
    return data

def dataframe_below_marker_or_first(xlsx_path: str, marker="liste des resultats"):
    wb = load_workbook(xlsx_path, data_only=True)
    marker_norm = normalize_colname(marker)
    found_sheet, found_row = None, None
    for ws in wb.worksheets:
        for r in ws.iter_rows():
            for cell in r:
                v = cell.value
                if v is None: continue
                if normalize_colname(str(v)) == marker_norm:
                    found_sheet, found_row = ws.title, cell.row
                    break
            if found_sheet: break
        if found_sheet: break
    if found_sheet is None:
        return read_after_skip(xlsx_path, 0)
    df_all = pd.read_excel(xlsx_path, sheet_name=found_sheet, header=None, engine="openpyxl")
    df = df_all.iloc[found_row:].copy()
    header_rel_idx = None
    for i in range(len(df)):
        row = df.iloc[i]
        if row.notna().sum() >= 2:
            header_rel_idx = i
            break
    if header_rel_idx is None:
        return pd.DataFrame()
    header_row = df.iloc[header_rel_idx].astype(str).str.strip().tolist()
    data = df.iloc[header_rel_idx+1:].copy()
    data.columns = header_row
    data = data.loc[:, data.columns.notnull()]
    data = data.dropna(how='all')
    return data

def to_decimal(value) -> Decimal:
    """Convertit divers formats ('1 234,56€', float, int) en Decimal. NaN -> 0."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return Decimal('0')
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    s = str(value).strip()
    s = s.replace('\u00a0', '').replace('\u202f', '')  # espaces insécables
    s = s.replace('€', '').replace(' ', '')
    s = s.replace(',', '.')  # décimale FR -> point
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return Decimal('0')

def pick_column(existing_cols, synonyms):
    norm_map = {normalize_colname(c): c for c in existing_cols}
    for syn in synonyms:
        if syn in norm_map: return norm_map[syn]
    for syn in synonyms:
        for norm, original in norm_map.items():
            if syn in norm: return original
    return None

def sig_value(x):
    """Valeur canonique pour signature de ligne (déduplication stricte)."""
    if isinstance(x, (pd.Timestamp, dt.datetime, dt.date)):
        return date_to_text_dmy(x)
    if isinstance(x, Decimal):
        return format(x, "f")
    if isinstance(x, float):
        return format(x, ".10g")
    return "" if x is None else str(x)

# -------- Synonymes --------
SYN = {
    "N° commande": [
        "n commande","no commande","numero commande","n de commande","n commande","n° commande",
        "num commande","n cmd","no cmd","numero cmd","cmd","commande","order","order id","bdc"
    ],
    "Libellé": ["libelle","désignation","designation","objet","description","intitule","intitulé","libellé","objet"],
    "Fournisseur": ["fournisseur","vendor","tiers","fournisseu"],
    "Montant HT": ["montant ht","total ht","ht","montant hors taxes","m ht","mnt ht","montantht"],
    "Ind. Visa": ["ind visa","indice visa","indicateur visa","visa","visa ind","visa (ind)","ind? visa","ind.? visa"],
    "Statut": ["statut","status","etat","état"],
    "Nature de dépense": ["nature de depense","nature de dépense","nature depense","nature dépense","nature de la depense","nature de la dépense","type de depense","type de dépense","nature"],
    "Type de flux": ["type de flux","flux","nature de flux"],
    "Auteur": ["auteur","saisi par","cree par","créé par"],
    "Date de règlement": ["date de reglement","date reglement","date de paiement","date paiement","reglement","paiement"],
    "Commande": ["commande","n commande","no commande","numero commande","n° commande","cmd","bdc"],
    "Statut (constatations)": ["statut","etat","état"],
    "Date": ["date","date workflow","workflow","date de workflow","dt workflow","maj","mise a jour","mise à jour"]
}

# -------- Process --------
def process_commandes(path: str) -> pd.DataFrame:
    df = read_after_skip(path, 20)
    cols = list(df.columns)
    def col(name): return pick_column(cols, SYN[name])
    # filtres
    c_f = col("Fournisseur"); c_n = col("Nature de dépense")
    if c_f is not None:
        df = df[~df[c_f].astype(str).str.strip().str.upper().eq("FCM 3MUNDI ESR-M")]
    if c_n is not None:
        nature_clean = df[c_n].astype(str).map(strip_accents).str.lower().str.strip()
        df = df[nature_clean != "mission"]
    # ordre final
    order_map = [
        ("N° commande","N° commande"),
        ("Libellé","Libellé"),
        ("Fournisseur","Fournisseur"),
        ("Montant HT","Montant HT"),
        ("Type de flux","Type de flux"),
        ("Nature de dépense","Nature de dépense"),
        ("Statut","Statut"),
        ("Ind. Visa","Ind. Visa"),
        ("Auteur","Auteur"),
    ]
    out = pd.DataFrame()
    for target, synkey in order_map:
        c = col(synkey); out[target] = df[c] if c is not None else None
    return out.dropna(how="all")

def process_constatations(path: str) -> pd.DataFrame:
    df = read_after_skip(path, 17)
    cols = list(df.columns)
    c_cmd = pick_column(cols, SYN["Commande"])
    c_stat = pick_column(cols, SYN["Statut (constatations)"])
    out = pd.DataFrame()
    out["Commande"] = df[c_cmd] if c_cmd else None
    out["extrait commande"] = df[c_cmd].astype(str).str.slice(0,5) if c_cmd else None
    out["Statut"] = df[c_stat] if c_stat else None
    return out.dropna(how="all")

def process_envoi_bdc(path: str) -> pd.DataFrame:
    df = read_after_skip(path, 0)
    df = df.iloc[:, :3].copy()
    while df.shape[1] < 3:
        df[df.shape[1]] = None
    df.columns = ["Commande", "Date envoi", "Agent"]
    return df.dropna(how="all")

def process_factures(path: str) -> pd.DataFrame:
    df = read_after_skip(path, 19)
    cols = list(df.columns)
    c_nat = pick_column(cols, SYN["Nature de dépense"])
    c_fou = pick_column(cols, SYN["Fournisseur"])
    if c_nat is not None:
        nat_clean = df[c_nat].astype(str).str.strip().str.upper()
        df = df[~nat_clean.eq("MI")]
    if c_fou is not None:
        df = df[~df[c_fou].astype(str).str.strip().str.upper().eq("FCM 3MUNDI ESR-M")]
    c_bdc = pick_column(cols, SYN["N° commande"])
    c_ht  = pick_column(cols, SYN["Montant HT"])
    c_reg = pick_column(cols, SYN["Date de règlement"])
    out = pd.DataFrame()
    out["N° commande"] = df[c_bdc] if c_bdc else None
    out["Montant HT"] = df[c_ht] if c_ht else None
    out["Date de règlement"] = df[c_reg] if c_reg else None
    return out.dropna(how="all")

def process_workflow(path: str) -> pd.DataFrame:
    return dataframe_below_marker_or_first(path, marker="Liste des résultats")

# -------- Auto-fit --------
def autofit_worksheet(ws, df: pd.DataFrame, min_width=10, max_width=60, padding=2):
    for idx, col_name in enumerate(df.columns, start=1):
        col_vals = df[col_name].astype(str).fillna("")
        max_len = max([len(str(col_name))] + [len(v) for v in col_vals])
        ws.column_dimensions[get_column_letter(idx)].width = max(min_width, min(max_len + padding, max_width))

# -------- Global --------
def choose_workflow_value_column(df_wf: pd.DataFrame):
    if df_wf is None or df_wf.empty:
        return None, None
    cols = list(df_wf.columns)
    c_bdc = pick_column(cols, SYN["N° commande"])
    for key in ("Date", "Statut"):
        c = pick_column(cols, SYN[key]) if key in SYN else None
        if c: return c_bdc, c
    if len(cols) >= 2:
        return c_bdc, cols[1]
    return c_bdc, None

def create_and_fill_global_sheet(writer, df_cmd, df_envoi, df_fact, df_wf, df_const):
    book = writer.book
    ws = book.create_sheet("Global")

    ws.page_setup.orientation = 'landscape'
    ws.page_margins.left = 0.19685  # 0,5 cm
    ws.page_margins.right = 0.19685

    headers = ["BDC", "OBJET", "FOURN.", "HT", "VISA", "ENVOYE", "SF", "WORKFLOW", "PAYE", "SOLDE", "STATUT"]
    ws.append(headers)

    header_font = Font(name="Calibri", size=12)
    header_align = Alignment(horizontal="center", vertical="center")
    for c in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=c); cell.font = header_font; cell.alignment = header_align

    for i, w in enumerate(GLOBAL_COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w + GLOBAL_WIDTH_OFFSET

    body_font = Font(name="Calibri", size=9)
    body_align = Alignment(horizontal="center", vertical="center")

    if df_cmd is None or df_cmd.empty or "N° commande" not in df_cmd.columns:
        ws.row_dimensions[1].height = 30
        return ws

    # Envoi BDC -> F
    envoi_lookup = {}
    if df_envoi is not None and not df_envoi.empty:
        for _, r in df_envoi.iterrows():
            key = str(r.get("Commande", "")).strip()
            if not key: continue
            b_txt = date_to_text_dmy(r.get("Date envoi"))
            c_txt = "" if pd.isna(r.get("Agent")) else str(r.get("Agent")).strip()
            val = f"{b_txt} {c_txt}".strip() if (b_txt or c_txt) else ""
            if key not in envoi_lookup:
                envoi_lookup[key] = val

    # Factures -> agrégats pour I (Payé) et J (Solde)
    # Stocke date parsée ET valeur brute pour le cas 1 facture sans date exploitable
    fact_agg = {}  # bdc -> {'count': n, 'sum': Decimal, 'date': dt.date|None, 'raw': Any|None}
    if df_fact is not None and not df_fact.empty and "N° commande" in df_fact.columns:
        for _, r in df_fact.iterrows():
            key = str(r.get("N° commande", "")).strip()
            if not key: continue
            amt = to_decimal(r.get("Montant HT"))
            raw_date = r.get("Date de règlement")
            dreg = to_date_only(raw_date)  # dt.date | "" | original si non parsé

            if key not in fact_agg:
                fact_agg[key] = {'count': 0, 'sum': Decimal('0'), 'date': None, 'raw': None}

            fact_agg[key]['count'] += 1
            fact_agg[key]['sum'] += amt

            if fact_agg[key]['count'] == 1:
                fact_agg[key]['date'] = dreg if isinstance(dreg, dt.date) else None
                fact_agg[key]['raw']  = raw_date
            else:
                fact_agg[key]['date'] = None
                fact_agg[key]['raw']  = None

    # Workflow -> H
    wf_lookup = {}
    wf_bdc_col, wf_val_col = choose_workflow_value_column(df_wf)
    if df_wf is not None and not df_wf.empty and wf_bdc_col is not None:
        for _, r in df_wf.iterrows():
            key = str(r.get(wf_bdc_col, "")).strip()
            if not key: continue
            v = r.get(wf_val_col) if wf_val_col else ""
            v = to_date_only(v)
            wf_lookup[key] = v

    # Constatation -> Statut (pour G)
    const_stat_by_full, const_stat_by_extract = {}, {}
    if df_const is not None and not df_const.empty:
        for _, r in df_const.iterrows():
            key_full = str(r.get("Commande", "")).strip()
            key_ex = str(r.get("extrait commande", "")).strip()
            st = r.get("Statut")
            if key_full: const_stat_by_full[key_full] = st
            if key_ex: const_stat_by_extract[key_ex] = st

    # Déduplication stricte : signature (toutes colonnes)
    seen_signatures = set()

    # Lignes Global
    for _, row in df_cmd.iterrows():
        bdc = str(row.get("N° commande", "")).strip()
        if not bdc: continue

        b = row.get("Libellé", "-")
        c = row.get("Fournisseur", "-")
        d_val_raw = row.get("Montant HT", "0")
        d = d_val_raw
        e = row.get("Ind. Visa", "-")
        f = envoi_lookup.get(bdc, "")

        # G (SF)
        g = "Pas de SF connu"
        c_norm = str(c).strip().upper()
        if c_norm == "BNP PARIBAS - REGULARISATION CARTE ACHAT":
            g = "ss objet Régul CA"
        else:
            f_norm = strip_accents(str(f)).lower()
            if "ss objet regul ca" in f_norm:
                st = const_stat_by_full.get(bdc)
                if st is None or (isinstance(st, float) and pd.isna(st)):
                    st = const_stat_by_extract.get(bdc[:5])
                g = st if st not in (None, "") and not (isinstance(st, float) and pd.isna(st)) else "Pas de SF connu"

        # H (WORKFLOW)
        h = wf_lookup.get(bdc, "")

        # I (PAYE)
        fa = fact_agg.get(bdc)
        if fa is None or fa['count'] == 0:
            i = "pas de paiement connu"
        elif fa['count'] == 1:
            if isinstance(fa['date'], dt.date):
                i = fa['date']  # vraie date
            else:
                raw = fa['raw']
                txt = date_to_text_dmy(raw) if raw not in (None, "") else "date manquante"
                i = txt
        else:
            n = fa['count']
            i = f"{n} paiement" + ("s" if n >= 2 else "")

        # J (SOLDE) = D - somme(Montant HT factures)
        total_fact = fact_agg.get(bdc, {'sum': Decimal('0')})['sum']
        d_amount = to_decimal(d_val_raw)
        solde = d_amount - total_fact
        j = float(solde)

        # K (STATUT)
        k = row.get("Statut", "-")

        row_values = [bdc, b, c, d, e, f, g, h, i, j, k]
        signature = tuple(sig_value(x) for x in row_values)
        if signature in seen_signatures:
            continue  # doublon strict -> on ignore
        seen_signatures.add(signature)

        ws.append(row_values)

    # Mise en forme corps
    max_row = ws.max_row
    for r in range(2, max_row+1):
        ws.cell(row=r, column=1).number_format = "@"
        for col in (8, 9):  # H, I si dates
            cell = ws.cell(row=r, column=col)
            if isinstance(cell.value, (pd.Timestamp, dt.datetime, dt.date)):
                cell.number_format = "dd/mm/yyyy"
        jcell = ws.cell(row=r, column=10)  # J solde
        if isinstance(jcell.value, (int, float)):
            jcell.number_format = "0.00"
        ws.row_dimensions[r].height = 30
        for ccol in range(1, len(headers)+1):
            cell = ws.cell(row=r, column=ccol); cell.font = body_font; cell.alignment = body_align

    ws.row_dimensions[1].height = 30
    return ws

def create_cover_sheet(writer):
    book = writer.book
    ws = book.create_sheet("Page de garde")
    ws["A1"] = GLOBAL_COVER_TEXT
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 120
    ws.row_dimensions[1].height = 600
    return ws

# -------- GUI --------
BaseTk = TkinterDnD.Tk if DND_AVAILABLE else tk.Tk

class App(BaseTk):
    def __init__(self):
        super().__init__()
        self.title("Nettoie XLSX (Commandes / Constatations / Factures / Envoi BDC / Workflow)")
        self.geometry("980x620")
        self.resizable(False, False)

        padd = {'padx': 8, 'pady': 6}
        frm = ttk.Frame(self); frm.pack(fill="both", expand=True, **padd)

        # Variables
        self.commandes_var = tk.StringVar()
        self.constatations_var = tk.StringVar()
        self.factures_var = tk.StringVar()
        self.envoi_var = tk.StringVar()
        self.workflow_var = tk.StringVar()
        self.outfile_var = tk.StringVar()

        # Lignes fichiers (ordre demandé)
        self._row_file(frm, "Commandes (.xlsx)", self.commandes_var)
        self._row_file(frm, "Constatations (.xlsx)", self.constatations_var)
        self._row_file(frm, "Factures (.xlsx)", self.factures_var)
        self._row_file(frm, "Envoi BDC (.xlsx)", self.envoi_var)
        self._row_file(frm, "Workflow (.xlsx)", self.workflow_var)

        # Sortie
        ttk.Label(frm, text="Fichier de sortie (.xlsx)").grid(row=5, column=0, sticky="w")
        ttk.Entry(frm, textvariable=self.outfile_var, width=78).grid(row=5, column=1, sticky="we")
        ttk.Button(frm, text="Parcourir…", command=self._pick_outfile).grid(row=5, column=2, sticky="we")

        # Boutons
        btns = ttk.Frame(frm); btns.grid(row=6, column=0, columnspan=3, sticky="we", pady=10)
        ttk.Button(btns, text="Lancer le traitement", command=self.run).pack(side="left", padx=4)
        ttk.Button(btns, text="Vider les champs", command=self.clear_fields).pack(side="left", padx=4)
        ttk.Button(btns, text="Règles Global (A→K)", command=self.show_rules).pack(side="left", padx=12)

        # Log (exactement tes lignes)
        self.log = tk.Text(frm, height=16)
        self.log.grid(row=7, column=0, columnspan=3, sticky="nsew", pady=(6,0))
        frm.grid_columnconfigure(1, weight=1)

        for line in INTRO_LOG_TEXT.strip().splitlines():
            self._log(line)
        if not DND_AVAILABLE:
            self._log("Drag & drop indisponible : installez 'tkinterdnd2' (pip install tkinterdnd2).")

    # UI helpers
    def _row_file(self, parent, label, var):
        r = len(parent.grid_slaves()) // 3
        ttk.Label(parent, text=label).grid(row=r, column=0, sticky="w")
        entry = ttk.Entry(parent, textvariable=var, width=78); entry.grid(row=r, column=1, sticky="we")
        ttk.Button(parent, text="Parcourir…", command=lambda v=var: self._pick(v)).grid(row=r, column=2, sticky="we")
        if DND_AVAILABLE:
            entry.drop_target_register(DND_FILES)
            entry.dnd_bind('<<Drop>>', lambda e, v=var: self._on_drop(e, v))

    def _pick(self, var):
        path = filedialog.askopenfilename(filetypes=[("Excel","*.xlsx")])
        if path: var.set(path)

    def _pick_outfile(self):
        initial_dir = ""
        for v in (self.commandes_var.get(), self.constatations_var.get(), self.factures_var.get(),
                  self.envoi_var.get(), self.workflow_var.get()):
            if v: initial_dir = os.path.dirname(v); break
        path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                            filetypes=[("Excel","*.xlsx")],
                                            initialdir=initial_dir or None,
                                            initialfile="export_clean.xlsx")
        if path: self.outfile_var.set(path)

    def _on_drop(self, event, var):
        try:
            paths = self.tk.splitlist(event.data)
            if not paths: return
            for p in paths:
                if p.lower().endswith(".xlsx"):
                    var.set(p); break
            else:
                messagebox.showwarning("Format non pris en charge", "Déposez un fichier .xlsx.")
        except Exception as e:
            messagebox.showerror("Erreur DnD", str(e))

    def clear_fields(self):
        self.commandes_var.set(""); self.constatations_var.set("")
        self.factures_var.set(""); self.envoi_var.set("")
        self.workflow_var.set(""); self.outfile_var.set("")
        self.log.delete("1.0","end")

    def show_rules(self):
        win = tk.Toplevel(self)
        win.title("Règles Global (A→K)")
        win.geometry("720x520")
        win.resizable(True, True)
        txt = tk.Text(win, wrap="word")
        txt.pack(fill="both", expand=True)
        txt.insert("1.0", GLOBAL_RULES_TEXT)
        txt.config(state="disabled")

    def _log(self, msg):
        self.log.insert("end", msg+"\n"); self.log.see("end"); self.update_idletasks()

    # Run
    def run(self):
        files = {
            "Commandes": self.commandes_var.get().strip(),
            "Constatations": self.constatations_var.get().strip(),
            "Factures": self.factures_var.get().strip(),
            "EnvoiBDC": self.envoi_var.get().strip(),
            "Workflow": self.workflow_var.get().strip(),
        }
        if not any(files.values()):
            messagebox.showwarning("Aucun fichier","Sélectionnez au moins un fichier à traiter."); return

        outfile = self.outfile_var.get().strip()
        if not outfile:
            self._pick_outfile(); outfile = self.outfile_var.get().strip()
            if not outfile:
                messagebox.showwarning("Sortie manquante","Veuillez choisir un fichier de sortie .xlsx."); return

        try:
            dfs = {}

            if files["Commandes"]:
                self._log("Lecture/Nettoyage : Commandes")
                dfs["Commande"] = process_commandes(files["Commandes"])

            if files["Constatations"]:
                self._log("Lecture/Nettoyage : Constatations")
                dfs["Constatation"] = process_constatations(files["Constatations"])

            if files["Factures"]:
                self._log("Lecture/Nettoyage : Factures")
                dfs["Factures"] = process_factures(files["Factures"])

            if files["EnvoiBDC"]:
                self._log("Lecture/Nettoyage : Envoi BDC")
                dfs["Envoi BDC"] = process_envoi_bdc(files["EnvoiBDC"])

            if files["Workflow"]:
                self._log("Lecture : Workflow")
                dfs["Workflow"] = process_workflow(files["Workflow"])

            # Écriture des feuilles sources
            order = ["Commande", "Envoi BDC", "Constatation", "Factures", "Workflow"]
            with pd.ExcelWriter(outfile, engine="openpyxl") as writer:
                self._log("Création de la page de garde")
                create_cover_sheet(writer)
                for name in order:
                    if name in dfs:
                        self._log(f"Écriture de la feuille {name}")
                        df = dfs[name]
                        df.to_excel(writer, index=False, sheet_name=name)
                        ws = writer.book[name]
                        autofit_worksheet(ws, df)
                        strip_times_in_worksheet(ws)

                # Global
                self._log("Création et remplissage de la feuille Global")
                create_and_fill_global_sheet(
                    writer,
                    dfs.get("Commande"),
                    dfs.get("Envoi BDC"),
                    dfs.get("Factures"),
                    dfs.get("Workflow"),
                    dfs.get("Constatation"),
                )

            self._log(f"✔ Terminé. Fichier créé : {outfile}")
            messagebox.showinfo("Terminé", f"Fichier créé :\n{outfile}")

        except Exception as e:
            self._log(f"✖ Erreur : {e}")
            messagebox.showerror("Erreur", f"Echec du traitement : {e}")

if __name__ == "__main__":
    app = App()
    app.mainloop()
